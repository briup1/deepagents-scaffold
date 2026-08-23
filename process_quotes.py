"""
处理 simple_quote.xlsx：
1. 生成凸显名字（起运港/目的港/船公司/柜型）的高亮版本。
2. 生成一份全新的、工整清晰的国际海运报价单（3条）。
"""

from __future__ import annotations

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def style_header(ws, header_fill, header_font):
    """表头统一样式。"""
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )


def style_data_cell(cell, *, bold: bool = False, color: str | None = None, wrap: bool = True):
    """数据单元格统一样式。"""
    font_kwargs = {"name": "Microsoft YaHei", "size": 11}
    if bold:
        font_kwargs["bold"] = True
    if color:
        font_kwargs["color"] = color
    cell.font = Font(**font_kwargs)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


def auto_column_width(ws, max_width: int = 40):
    """根据内容自动调整列宽。"""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    text = str(cell.value)
                    # 按换行取最长行
                    longest_line = max((len(line) for line in text.split("\n")), default=0)
                    max_len = max(max_len, longest_line)
        # 给点边距，上限 max_width
        width = min(max_len + 4, max_width)
        ws.column_dimensions[col_letter].width = width


def highlight_original():
    """任务一：在原表基础上凸显名字列。"""
    df = pd.read_excel("simple_quote.xlsx", sheet_name=0)

    wb = Workbook()
    ws = wb.active
    ws.title = "海运报价单（名字凸显）"

    # 写表头
    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)

    # 写数据
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 样式
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    style_header(ws, header_fill, header_font)

    # 名字列：报价单号、起运港、目的港、船公司、柜型（这些是海运报价中最核心的识别名）
    name_columns = {"报价单号", "起运港", "目的港", "船公司", "柜型"}
    name_col_idx = {col: i + 1 for i, col in enumerate(df.columns) if col in name_columns}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            is_name = cell.column in name_col_idx
            style_data_cell(cell, bold=is_name, color="0000FF" if is_name else None)

    # 行高
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A2"
    auto_column_width(ws)

    out_path = "simple_quote_highlighted.xlsx"
    wb.save(out_path)
    print(f"已生成：{out_path}")


def create_clean_quotes():
    """任务二：生成新的工整国际海运报价单（3条示例数据）。"""
    columns = [
        "报价单号", "起运港", "目的港", "中转港", "船公司", "船名航次",
        "柜型", "海运费 (USD)", "附加费 (USD)", "ALL-IN 小计 (USD)",
        "有效期", "ETD", "ETA", "航程 (天)", "截关 / 开船", "备注",
    ]

    rows = [
        [
            "Q-2025-091",
            "盐田 YANTIAN (CNYTN)",
            "汉堡 Hamburg (DEHAM)",
            "新加坡 Singapore",
            "MSC 地中海航运",
            "MSC GULSUN V.025E",
            "40'HC",
            1850,
            "BAF 285 + PSS 180",
            2315,
            "2025-09-15 ~ 2025-10-15",
            "2025-09-08",
            "2025-10-04",
            26,
            "周三截关 / 周五开船",
            "含 THC/ORC；不含目的港 DTHC；超重费 USD60/20'",
        ],
        [
            "Q-2025-092",
            "上海 SHANGHAI (CNSHA)",
            "洛杉矶 Los Angeles (USLAX)",
            "直航",
            "COSCO 中远海运",
            "COSCO SHANGHAI V.168E",
            "20'GP",
            1420,
            "FAF 95 + EBS 120",
            1635,
            "2025-09-20 ~ 2025-10-10",
            "2025-09-12",
            "2025-09-26",
            14,
            "周一截关 / 周三开船",
            "含启运港 THC；美西普船；限重 22 MT",
        ],
        [
            "Q-2025-093",
            "宁波 NINGBO (CNNGB)",
            "悉尼 Sydney (AUSYD)",
            "香港 Hong Kong",
            "ONE 海洋网联船务",
            "ONE COLUMBA V.073W",
            "40'HC",
            2650,
            "DOC 50 + BAF 220",
            2920,
            "2025-09-18 ~ 2025-10-08",
            "2025-09-16",
            "2025-10-05",
            19,
            "周四截关 / 周日开船",
            "含 ORC；澳洲线旺季，建议提前 7 天订舱",
        ],
    ]

    df = pd.DataFrame(rows, columns=columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "国际海运报价单"

    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    style_header(ws, header_fill, header_font)

    name_columns = {"起运港", "目的港", "中转港", "船公司", "船名航次", "柜型"}
    name_col_idx = {col: i + 1 for i, col in enumerate(df.columns) if col in name_columns}
    numeric_columns = {"海运费 (USD)", "ALL-IN 小计 (USD)", "航程 (天)"}
    numeric_col_idx = {col: i + 1 for i, col in enumerate(df.columns) if col in numeric_columns}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            is_name = cell.column in name_col_idx
            is_numeric = cell.column in numeric_col_idx
            color = "0000FF" if is_name else None
            style_data_cell(cell, bold=is_name, color=color)
            if is_numeric and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # 让价格行更醒目：ALL-IN 小计加浅黄底色
    all_in_idx = df.columns.get_loc("ALL-IN 小计 (USD)") + 1
    for r_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=r_idx, column=all_in_idx)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.row_dimensions[1].height = 28
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42

    ws.freeze_panes = "A2"
    auto_column_width(ws, max_width=36)

    out_path = "international_shipping_quote.xlsx"
    wb.save(out_path)
    print(f"已生成：{out_path}")


if __name__ == "__main__":
    highlight_original()
    create_clean_quotes()
    print("全部完成。")

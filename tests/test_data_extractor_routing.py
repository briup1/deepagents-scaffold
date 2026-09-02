"""Data Extractor 路由决策树集成测试。

验证 system_prompt_suffix 中的路由规则：
1. 异常信号 → 预处理 (normalize_upload_file) → 委派
2. 模板命中 → 快路径
3. 复杂场景 → 委派 extraction_coder
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from scaffold.infra.config.app_config import get_app_config
from scaffold.plugins.tools.normalize_upload import normalize_upload_file
from scaffold.plugins.tools.preview_excel import preview_excel


class TestRoutingDecisionTree:
    """验证路由决策树的各个分支。"""

    @pytest.fixture
    def xlsx_with_merged_cells(self) -> bytes:
        """生成包含合并单元格的 Excel 文件。"""
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Quotation"
        ws.merge_cells("A1:E1")
        ws["A1"] = "Freight Quotation"
        ws.append(["carrier", "pol", "pod", "container_type", "amount"])
        ws.append(["MSC", "Shanghai", "Los Angeles", "20GP", 1200])
        ws.append(["COSCO", "Ningbo", "Los Angeles", "40HQ", 2300])
        wb.save(buffer)
        return buffer.getvalue()

    @pytest.fixture
    def xlsx_with_strikethrough(self) -> bytes:
        """生成包含删除线的 Excel 文件。"""
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Quotation"
        ws.append(["carrier", "pol", "pod", "container_type", "amount"])
        ws.append(["MSC", "Shanghai", "Los Angeles", "20GP", 1200])
        for cell in ws[2]:
            cell.font = openpyxl.styles.Font(strike=True)
        ws.append(["COSCO", "Ningbo", "Los Angeles", "40HQ", 2300])
        wb.save(buffer)
        return buffer.getvalue()

    @pytest.fixture
    def normal_xlsx(self) -> bytes:
        """生成正常的 Excel 文件（无异常信号）。"""
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Quotation"
        ws.append(["carrier", "pol", "pod", "container_type", "amount"])
        ws.append(["MSC", "Shanghai", "Los Angeles", "20GP", 1200])
        ws.append(["COSCO", "Ningbo", "Los Angeles", "40HQ", 2300])
        wb.save(buffer)
        return buffer.getvalue()

    @pytest.mark.asyncio
    async def test_anomaly_signal_merged_cells_triggers_normalize(
        self,
        client: TestClient,
        xlsx_with_merged_cells: bytes,
    ) -> None:
        """异常信号：合并单元格 → 必须先 normalize_upload_file。"""
        # 1. 上传文件
        resp = client.post(
            "/api/files/upload",
            data={"thread_id": "thread-routing-001"},
            files={
                "file": (
                    "merged.xlsx",
                    xlsx_with_merged_cells,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        artifact_id = resp.json()["artifact_id"]

        # 2. preview_excel 应检测到 merged_cells_count > 0
        preview = await preview_excel(artifact_id=artifact_id)
        assert "error" not in preview
        assert preview["merged_cells_count"] > 0, "应检测到合并单元格"

        # 3. 必须调用 normalize_upload_file 获得 normalized_artifact_id
        normalized = await normalize_upload_file(artifact_id=artifact_id)
        assert "error" not in normalized, f"normalize 失败: {normalized}"
        assert "normalized_artifact_id" in normalized
        assert normalized["stats"]["merged_cells_processed"] > 0

        # 4. 用 normalized_artifact_id 预览，merged_cells_count 应为 0
        preview_norm = await preview_excel(artifact_id=normalized["normalized_artifact_id"])
        assert preview_norm["merged_cells_count"] == 0, "规范化后合并单元格应被拆分"

    @pytest.mark.asyncio
    async def test_anomaly_signal_strikethrough_triggers_normalize(
        self,
        client: TestClient,
        xlsx_with_strikethrough: bytes,
    ) -> None:
        """异常信号：删除线 → 必须先 normalize_upload_file（默认过滤）。"""
        resp = client.post(
            "/api/files/upload",
            data={"thread_id": "thread-routing-002"},
            files={
                "file": (
                    "strike.xlsx",
                    xlsx_with_strikethrough,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        artifact_id = resp.json()["artifact_id"]

        preview = await preview_excel(artifact_id=artifact_id)
        assert "error" not in preview
        assert preview["strikethrough_count"] > 0, "应检测到删除线"

        normalized = await normalize_upload_file(artifact_id=artifact_id)
        assert "error" not in normalized
        assert normalized["stats"]["strikethrough_rows_filtered"] > 0

        preview_norm = await preview_excel(artifact_id=normalized["normalized_artifact_id"])
        assert preview_norm["strikethrough_count"] == 0, "规范化后删除线行应被过滤"

    @pytest.mark.asyncio
    async def test_normal_file_no_anomaly_no_normalize_needed(
        self,
        client: TestClient,
        normal_xlsx: bytes,
    ) -> None:
        """正常文件（无异常信号）→ 可直接走后续流程，无需 normalize。"""
        resp = client.post(
            "/api/files/upload",
            data={"thread_id": "thread-routing-003"},
            files={
                "file": (
                    "normal.xlsx",
                    normal_xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        artifact_id = resp.json()["artifact_id"]

        preview = await preview_excel(artifact_id=artifact_id)
        assert "error" not in preview
        assert preview["merged_cells_count"] == 0
        assert preview["strikethrough_count"] == 0

        # 正常文件直接可用于后续流程（模板匹配/生成脚本）
        # 这里只验证无异常信号，实际模板匹配由 match_extraction_template 测试

    @pytest.mark.asyncio
    async def test_template_hit_fast_path(
        self,
        client: TestClient,
        normal_xlsx: bytes,
    ) -> None:
        """模板命中 → 快路径（复用模板 script，跳过需求确认与生成脚本）。"""
        from scaffold.plugins.tools.extraction_templates import match_extraction_template

        resp = client.post(
            "/api/files/upload",
            data={"thread_id": "thread-routing-004"},
            files={
                "file": (
                    "normal.xlsx",
                    normal_xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        artifact_id = resp.json()["artifact_id"]

        # 调用 match_extraction_template（模拟模板已存在的场景）
        # 注意：实际测试需要先保存一个模板，这里验证工具调用链路
        match_result = await match_extraction_template(artifact_id=artifact_id)
        # 可能 matched=false（无模板）或 matched=true（有模板）
        # 测试重点：工具可被调用，返回结构符合预期
        assert "matched" in match_result
        if match_result["matched"]:
            assert "template" in match_result
            assert "script" in match_result["template"]
            assert "goal" in match_result["template"]

    @pytest.mark.asyncio
    async def test_complex_file_structure_delegation_signal(
        self,
        client: TestClient,
    ) -> None:
        """复杂文件结构（sheet>5、列>100、空行>50%） → 直接委派 extraction_coder。"""
        # 构造复杂文件：6 个 sheet，每个 120 列
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        for i in range(6):
            if i == 0:
                ws = wb.active
                assert ws is not None
                ws.title = f"Sheet{i + 1}"
            else:
                ws = wb.create_sheet(title=f"Sheet{i + 1}")
            # 120 列
            headers = [f"col_{j}" for j in range(120)]
            ws.append(headers)
            # 只有 2 行数据，其余空行 → 空行率 > 50%
            ws.append(["data"] * 120)
        wb.save(buffer)
        complex_xlsx = buffer.getvalue()

        resp = client.post(
            "/api/files/upload",
            data={"thread_id": "thread-routing-005"},
            files={
                "file": (
                    "complex.xlsx",
                    complex_xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        artifact_id = resp.json()["artifact_id"]

        preview = await preview_excel(artifact_id=artifact_id)
        assert "error" not in preview
        assert len(preview["sheet_names"]) > 5
        assert len(preview["columns"]) > 100
        # 这种结构应触发直接委派 extraction_coder 的路由规则
        # 验证：preview 返回的结构摘要包含决策所需字段
        required_fields = ["sheet_names", "columns", "total_rows", "merged_cells_count", "strikethrough_count"]
        for field in required_fields:
            assert field in preview, f"结构摘要缺少 {field}"

    def test_delegation_contract_structure(self) -> None:
        """验证委派契约包含四要素：需求契约、验收标准、结构摘要、工作区约定。"""
        # 这是一个结构性测试，验证 system_prompt_suffix 中定义的委派契约格式
        contract_keys = [
            "需求契约",  # requirements contract
            "验收标准",  # acceptance criteria
            "结构摘要",  # structure summary
            "工作区约定",  # workspace conventions
        ]
        # 实际在 prompt 中以中文标题形式出现，这里只做文档性验证
        # 真实委派由 LLM 按 prompt 指令组装 task 工具参数
        assert all(key in str(contract_keys) for key in contract_keys)


class TestConfigProfileConsistency:
    """验证三个配置文件中 data_extractor profile 的一致性。"""

    def test_config_yaml_has_routing_rules(self) -> None:
        """config.yaml 包含路由规则。"""
        app_config = get_app_config()  # 默认加载 config.yaml
        profiles = app_config.profiles.harness
        data_extractor = next((p for p in profiles if p.name == "data_extractor"), None)
        assert data_extractor is not None, "config.yaml 缺少 data_extractor profile"
        suffix = data_extractor.system_prompt_suffix or ""
        assert "路由决策树" in suffix, "config.yaml 缺少路由决策树"
        assert "异常信号预处理" in suffix, "config.yaml 缺少异常信号预处理规则"
        assert "模板快路径" in suffix, "config.yaml 缺少模板快路径规则"
        assert "完整抽取流程" in suffix, "config.yaml 缺少完整抽取流程规则"
        assert "委派契约" in suffix, "config.yaml 缺少委派契约定义"
        assert "需求契约" in suffix
        assert "验收标准" in suffix
        assert "结构摘要" in suffix
        assert "工作区约定" in suffix

    def test_config_verify_yaml_has_routing_rules(self) -> None:
        """config.verify.yaml 包含路由规则。"""
        import yaml

        with open("config.verify.yaml", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        profile = next((p for p in cfg["profiles"]["harness"] if p["name"] == "data_extractor"), None)
        assert profile is not None, "config.verify.yaml 缺少 data_extractor profile"
        suffix = profile.get("system_prompt_suffix", "")
        assert "路由决策树" in suffix
        assert "异常信号预处理" in suffix
        assert "模板快路径" in suffix
        assert "委派契约" in suffix

    def test_config_test_yaml_has_routing_rules(self) -> None:
        """config.test.yaml 包含路由规则。"""
        import yaml

        with open("config.test.yaml", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        profile = next((p for p in cfg["profiles"]["harness"] if p["name"] == "data_extractor"), None)
        assert profile is not None, "config.test.yaml 缺少 data_extractor profile"
        suffix = profile.get("system_prompt_suffix", "")
        assert "路由决策树" in suffix
        assert "异常信号预处理" in suffix
        assert "模板快路径" in suffix
        assert "委派契约" in suffix

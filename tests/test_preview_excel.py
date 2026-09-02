"""preview_excel 工具测试。"""

from __future__ import annotations

import io

import pytest
import pytest_asyncio
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.styles import Font

from scaffold.plugins.tools.preview_excel import preview_excel
from scaffold.infra.extraction import get_extraction_workspace


def _make_excel_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"
    ws.append(["carrier", "pol", "pod", "container_type", "amount"])
    ws.append(["MSC", "SHANGHAI", "LOS ANGELES", "40HQ", 3200])
    ws.append(["COSCO", "SHANGHAI", "LOS ANGELES", "20GP", 1800])
    ws.append(["ONE", "NINGBO", "LONG BEACH", "40HQ", 3100])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _make_excel_with_anomalies() -> bytes:
    """创建包含合并单元格和删除线的 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    ws.append(["A", "B", "C"])
    ws.append(["1", "2", "3"])
    ws.append(["4", "5", "6"])
    # 合并 A1:B1
    ws.merge_cells("A1:B1")
    # C2 设置删除线
    ws["C2"].font = Font(strikethrough=True)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _create_upload_artifact(client: TestClient, excel_bytes: bytes, thread_id: str) -> str:
    """通过 API 上传文件创建 upload 类型工件。"""
    response = client.post(
        "/api/files/upload",
        data={"thread_id": thread_id},
        files={
            "file": (
                "quote.xlsx",
                io.BytesIO(excel_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    return response.json()["artifact_id"]


async def _create_normalized_artifact(excel_bytes: bytes, thread_id: str) -> str:
    """通过 workspace 直接创建 normalized 类型工件。"""
    async with get_extraction_workspace() as ws:
        artifact = await ws.save_artifact(
            thread_id=thread_id,
            artifact_type="normalized",
            filename="normalized.xlsx",
            content=excel_bytes,
            original_name="normalized.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return artifact.artifact_id


async def _create_script_artifact() -> str:
    """通过 workspace 直接创建 script 类型工件。"""
    async with get_extraction_workspace() as ws:
        artifact = await ws.save_artifact(
            thread_id="t-reject",
            artifact_type="script",
            filename="script.py",
            content=b"print('hello')",
            original_name="script.py",
            mime_type="text/x-python",
        )
        return artifact.artifact_id


class TestPreviewExcel:
    @pytest.fixture
    def excel_bytes(self) -> bytes:
        return _make_excel_bytes()

    @pytest.fixture
    def anomaly_excel_bytes(self) -> bytes:
        return _make_excel_with_anomalies()

    @pytest.fixture
    def artifact_id(self, client: TestClient, excel_bytes: bytes) -> str:
        return _create_upload_artifact(client, excel_bytes, "t-preview")

    @pytest_asyncio.fixture
    async def anomaly_artifact_id(self, client: TestClient, anomaly_excel_bytes: bytes) -> str:
        return _create_upload_artifact(client, anomaly_excel_bytes, "t-anomaly")

    @pytest_asyncio.fixture
    async def normalized_artifact_id(self, excel_bytes: bytes) -> str:
        return await _create_normalized_artifact(excel_bytes, "t-normalized")

    @pytest_asyncio.fixture
    async def script_artifact_id(self) -> str:
        return await _create_script_artifact()

    async def test_preview_excel_success(self, artifact_id: str) -> None:
        result = await preview_excel(artifact_id=artifact_id, limit=2)

        assert "error" not in result
        assert result["sheet_names"] == ["Quotes"]
        assert result["columns"] == ["carrier", "pol", "pod", "container_type", "amount"]
        assert len(result["sample_rows"]) == 2
        assert result["total_rows"] == 3
        # 新增异常信号字段
        assert "merged_cells_count" in result
        assert "strikethrough_count" in result
        assert result["merged_cells_count"] == 0
        assert result["strikethrough_count"] == 0

    async def test_preview_excel_invalid_sheet(self, artifact_id: str) -> None:
        result = await preview_excel(artifact_id=artifact_id, sheet_index=5)
        assert "error" in result
        assert "超出范围" in result["error"]

    async def test_preview_excel_not_found(self) -> None:
        result = await preview_excel(artifact_id="art-nonexistent")
        assert "error" in result
        assert "不存在" in result["error"]

    async def test_preview_excel_anomaly_signals(self, anomaly_artifact_id: str) -> None:
        """测试异常信号：合并单元格和删除线计数。"""
        result = await preview_excel(artifact_id=anomaly_artifact_id)

        assert "error" not in result
        assert result["merged_cells_count"] == 1  # A1:B1 合并
        assert result["strikethrough_count"] == 1  # C2 有删除线

    async def test_preview_excel_normalized_artifact(self, normalized_artifact_id: str) -> None:
        """测试 normalized 类型工件也可预览。"""
        result = await preview_excel(artifact_id=normalized_artifact_id)

        assert "error" not in result
        assert result["sheet_names"] == ["Quotes"]
        assert result["columns"] == ["carrier", "pol", "pod", "container_type", "amount"]
        assert len(result["sample_rows"]) == 3
        assert result["total_rows"] == 3

    async def test_preview_excel_rejects_other_types(self, script_artifact_id: str) -> None:
        """测试非 upload/normalized 类型工件被拒绝。"""
        result = await preview_excel(artifact_id=script_artifact_id)
        assert "error" in result
        assert "不支持预览" in result["error"]

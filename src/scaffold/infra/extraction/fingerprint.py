"""Excel 结构指纹：复用 openpyxl 读取的 sheet 名与各 sheet 表头列计算。

指纹 JSON 形态（design.md 3.5）：

    {"sheets": ["报价单"], "columns": {"报价单": ["品名", "单价", "数量"]},
     "signature": "sha256(sheets+columns) 前 16 位"}

匹配规则：signature 完全一致才算候选（保守方向）。
"""

from __future__ import annotations

import hashlib
import json
from io import BytesIO

import openpyxl


def _canonical_structure(sheet_names: list[str], columns_by_sheet: dict[str, list[str]]) -> str:
    """将结构信息序列化为稳定字符串（键有序、排序稳定），作为签名输入。"""
    payload = {
        "sheets": sheet_names,
        "columns": {name: columns_by_sheet.get(name, []) for name in sheet_names},
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def compute_fingerprint(content: bytes) -> dict:
    """从 Excel 文件字节计算结构指纹。

    Returns:
        {"sheets": [...], "columns": {...}, "signature": "16 位 hex"}
    """
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    columns: dict[str, list[str]] = {}
    for name in sheet_names:
        ws = wb[name]
        row = next(ws.iter_rows(values_only=True), None)
        columns[name] = [str(cell) for cell in row] if row else []
    wb.close()

    canonical = _canonical_structure(sheet_names, columns)
    signature = hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]
    return {"sheets": sheet_names, "columns": columns, "signature": signature}

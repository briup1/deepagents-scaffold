"""上传文件规范化工具。

接收上传的 Excel 文件，按封装语义进行预处理：
- 合并单元格拆分：区域内所有单元格填充同值
- 删除线处理：默认过滤该行（config 可覆盖）
产出 normalized 类型工件，记录 source_upload_artifact_id。
"""

from __future__ import annotations

import asyncio
import logging
from io import BytesIO
from typing import Any

import openpyxl

from scaffold.infra.config.app_config import get_app_config
from scaffold.infra.extraction import get_extraction_workspace

logger = logging.getLogger(__name__)


async def normalize_upload_file(
    artifact_id: str,
    sheet_index: int = 0,
    filter_strikethrough: bool | None = None,
    **kwargs: Any,
) -> dict:
    """规范化上传的 Excel 文件。

    Args:
        artifact_id: 上传工件的 ID（artifact_type 必须为 "upload"）。
        sheet_index: 要处理的 sheet 索引，默认 0。
        filter_strikethrough: 是否过滤包含删除线的行。
            None 表示使用配置默认值（默认 True），True 过滤，False 保留。

    Returns:
        包含 normalized_artifact_id、源文件信息、处理统计的字典。
    """
    logger.info(
        "normalize_upload_file 被调用: artifact_id=%s sheet_index=%s filter_strikethrough=%s",
        artifact_id,
        sheet_index,
        filter_strikethrough,
    )

    async with get_extraction_workspace() as ws:
        artifact = await ws.get_artifact(artifact_id)
        if artifact is None:
            return {"error": f"工件 {artifact_id} 不存在"}

        if artifact.artifact_type != "upload":
            return {"error": f"工件 {artifact_id} 不是上传文件，artifact_type={artifact.artifact_type}"}

        try:
            content = await ws.read_artifact(artifact_id)
        except FileNotFoundError:
            return {"error": f"工件文件不存在：{artifact.stored_path}"}

        thread_id = artifact.thread_id

    app_config = get_app_config()
    if filter_strikethrough is None:
        filter_strikethrough = app_config.normalize.get("filter_strikethrough_default", True)

    def _process_excel() -> tuple[bytes, dict[str, Any]]:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=False)
        sheet_names = wb.sheetnames
        if sheet_index >= len(sheet_names):
            return b"", {
                "error": f"sheet_index {sheet_index} 超出范围，共有 {len(sheet_names)} 个 sheet",
                "sheet_names": sheet_names,
            }

        ws = wb.worksheets[sheet_index]

        merged_ranges = list(ws.merged_cells.ranges)
        merged_count = 0
        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            )
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_value
            merged_count += 1

        rows_to_delete = []
        strikethrough_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
            has_strikethrough = False
            for cell in row:
                if cell.font and cell.font.strike:
                    has_strikethrough = True
                    break
            if has_strikethrough:
                strikethrough_count += 1
                if filter_strikethrough:
                    rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        output = BytesIO()
        wb.save(output)
        normalized_bytes = output.getvalue()

        stats = {
            "merged_cells_processed": merged_count,
            "strikethrough_rows_found": strikethrough_count,
            "strikethrough_rows_filtered": len(rows_to_delete),
            "total_rows_after": ws.max_row,
        }
        return normalized_bytes, stats

    normalized_bytes, stats = await asyncio.to_thread(_process_excel)

    if "error" in stats:
        return stats

    async with get_extraction_workspace() as ws:
        filename = artifact.original_name or f"normalized_{artifact_id}.xlsx"
        base_name = filename.rsplit(".", 1)[0]
        ext = filename.rsplit(".", 1)[1] if "." in filename else "xlsx"
        normalized_filename = f"{base_name}_normalized.{ext}"

        normalized_artifact = await ws.save_artifact(
            thread_id=thread_id,
            artifact_type="normalized",
            filename=normalized_filename,
            content=normalized_bytes,
            original_name=normalized_filename,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={
                "source_upload_artifact_id": artifact_id,
                "source_original_name": artifact.original_name,
                "sheet_index": sheet_index,
                "filter_strikethrough": filter_strikethrough,
                **stats,
            },
        )

    return {
        "normalized_artifact_id": normalized_artifact.artifact_id,
        "source_upload_artifact_id": artifact_id,
        "normalized_filename": normalized_filename,
        "stats": stats,
    }

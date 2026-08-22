"""Excel 预览与结构抽取工具。

提供读取已上传 Excel 文件并返回结构信息的能力，供 Agent 在生成抽取脚本前预览文件内容。
"""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path

import aiosqlite
import openpyxl

from scaffold.infra.artifacts import ArtifactRepository, ArtifactStorage
from scaffold.infra.config.app_config import get_app_config


def _get_storage() -> ArtifactStorage:
    """根据当前配置获取工件存储实例。"""
    config = get_app_config()
    base_dir = Path(config.database.sqlite_dir or "./data") / "artifacts"
    return ArtifactStorage(base_dir)


async def _get_artifact(artifact_id: str):
    """根据 artifact_id 查询工件元数据。"""
    config = get_app_config()
    db_path = config.database.history_db or f"{config.database.sqlite_dir or './data'}/history.db"
    conn = await aiosqlite.connect(db_path)
    try:
        repo = ArtifactRepository(conn)
        return await repo.get(artifact_id)
    finally:
        await conn.close()


async def preview_excel(
    artifact_id: str,
    sheet_index: int = 0,
    limit: int = 20,
) -> dict:
    """预览已上传 Excel 文件的结构。

    Args:
        artifact_id: 上传工件的 ID。
        sheet_index: 要预览的 sheet 索引，默认 0。
        limit: 返回的最大样本行数（不含表头），默认 20。

    Returns:
        包含 sheet_names、columns、sample_rows、total_rows 的字典。
    """
    artifact = await _get_artifact(artifact_id)
    if artifact is None:
        return {"error": f"工件 {artifact_id} 不存在"}

    if artifact.artifact_type != "upload":
        return {"error": f"工件 {artifact_id} 不是上传文件"}

    storage = _get_storage()
    try:
        content = await asyncio.to_thread(storage.read, artifact.stored_path)
    except FileNotFoundError:
        return {"error": f"工件文件不存在：{artifact.stored_path}"}

    def _parse() -> dict:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet_names = wb.sheetnames
        if sheet_index >= len(sheet_names):
            return {
                "error": f"sheet_index {sheet_index} 超出范围，共有 {len(sheet_names)} 个 sheet",
                "sheet_names": sheet_names,
            }
        ws = wb.worksheets[sheet_index]
        rows = list(ws.iter_rows(values_only=True))
        columns = [str(cell) for cell in rows[0]] if rows else []
        sample_rows = rows[1 : limit + 1]
        total_rows = max(0, len(rows) - 1)
        return {
            "sheet_names": sheet_names,
            "columns": columns,
            "sample_rows": [list(row) for row in sample_rows],
            "total_rows": total_rows,
        }

    return await asyncio.to_thread(_parse)
